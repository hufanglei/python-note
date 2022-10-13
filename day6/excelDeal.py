from openpyxl import Workbook,load_workbook
# 创建excel
#
# #实例化
wb = Workbook()
#获取当前active的sheet
sheet = wb.active

print(sheet.title) #打印sheet的表名

sheet.title = "Alex大王的姑娘们"

sheet["B9"] = "black girl"
sheet["C9"] = "171,48,84"
sheet.append("Richel","178","49")

wb.save("excel_test.xlsx")

# 打开已有文件

# wb = load_workbook("excel_test.xlsx")


